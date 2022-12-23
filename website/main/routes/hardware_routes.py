
from flask import Blueprint, render_template, redirect, url_for, request, flash
from website.extensions import *
from ..forms import AddHardware, AddHardwareFromField, ReturnHardware
import openpyxl
from flask_login import login_required, current_user
from bson import ObjectId
from website.constants import *
from datetime import datetime

hardware = Blueprint('hardware', __name__)
collection = db_collection.find_one({"_id": "main"})
return_route = "/hardware/all"


def adjust_return_route(route, barcode):
    print(route)
    if route == 'details' or route == 'edit':
        return_route_id = db_items.find_one(
            {'barcode': barcode}, {'_id': 1})['_id']
        return_route = f"/hardware/show_info/present/{return_route_id}"
    else:
        return_route = "/hardware/all"
    return return_route


def update_history(db_key, db_value, modify_msg):
    data_for_history = db_items.find_one({db_key: db_value}, {
        '_id': 0
    })
    data_for_history['modyfikacja'] = modify_msg
    data_for_history['last_updated'] = datetime.now(
        local_tz).strftime("%Y-%m-%d %H:%M:%S")
    data_for_history['who_modified'] = current_user.login
    db_history.insert_one(data_for_history)
    update_for_cron("sm_history")


def go_through_file(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb['Sprzęt']
    data_table = []
    db_barcodes = []
    col_names = {}
    current_col = 0
    for col in ws.iter_cols(1, ws.max_column):
        col_names[col[0].value] = current_col
        current_col += 1
    local_time = datetime.now(local_tz).strftime("%Y-%m-%d %H:%M:%S")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row:
            def cell_data(cell):
                return cell if cell else ""
            barcode = row[col_names['Barcode']].value.strip()
            if check_existing_data(barcode, 'barcode', db_items):
                barcode = row[col_names['Barcode']].value.strip()
                stanowisko_data = cell_data(row[col_names['Stanowisko']].value)
                typ_data = cell_data(row[col_names['Typ']].value)
                marka_data = cell_data(row[col_names['Marka']].value)
                model_data = cell_data(row[col_names['Model']].value)
                system_data = cell_data(row[col_names['System']].value)
                mpk_data = cell_data(row[col_names['MPK']].value)
                opis_szkod = cell_data(
                    row[col_names['Opis uszkodzenia']].value)
                notatki = row[col_names['Uwagi']].value
                is_rented = row[col_names['Udostępniony']].value
                rent_date = row[col_names['Data udostępnienia']].value
                who_rented = row[col_names['Osoba udostępniająca']].value
                if type(rent_date) == str:
                    rent_date = datetime.strptime(rent_date, '"%Y-%m-%d')
                data = {
                    'stanowisko': data_handler(None, stanowisko_data, 'stanowisko'),
                    'typ': data_handler(None, typ_data, 'type'),
                    'marka': data_handler(None, marka_data, 'marka'),
                    'model': data_handler(None, model_data, 'model'),
                    'system': data_handler(None, system_data, 'system'),
                    'mpk': data_handler(None, mpk_data, 'mpk'),
                    'stan': row[col_names['Stan']].value if row[col_names['Stan']].value else "Sprawny",
                    'opis_uszkodzenia': opis_szkod,
                    'bitlocker': cell_data(row[col_names['hasło/bitlocker']].value),
                    'serial': cell_data(row[col_names['Nazwa / serial']].value),
                    'identyfikator': cell_data(row[col_names['Indentyfikator klucza odzyskiwania']].value),
                    'klucz_odzyskiwania': cell_data(row[col_names['Klucz/dysk odzyskiwania']].value),
                    'notatki': f"Sprzęt dodany z pliku - {datetime.now().strftime('%Y-%m-%d')}",
                    'rented_status': False,
                    'adder': current_user.login,
                    'upload_date': local_time,
                    'last_updated': local_time,
                }
                if notatki:
                    data['notatki'] = f"{notatki} ({data['notatki']})"
                if barcode:
                    data['barcode'] = barcode
                data_for_history = data.copy()
                data_for_history['modyfikacja'] = 'Stworzony'
                data_for_history['who_modified'] = current_user.login
                db_history.insert_one(data_for_history)
                data_for_history = None
                if is_rented:
                    projekt_data = cell_data(row[col_names['PROJEKT']].value)
                    lokalizacja_data = cell_data(
                        row[col_names['Lokalizacja']].value)
                    data['login'] = cell_data(row[col_names['Login']].value)
                    data['mocarz_id'] = cell_data(
                        row[col_names['Mocarz ID']].value)
                    data['projekt'] = data_handler(
                        None, projekt_data, 'projekt')
                    data['lokalizacja'] = data_handler(
                        None, lokalizacja_data, 'lokalizacja')
                    data['karta_zblizeniowa'] = cell_data(
                        row[col_names['Karta']].value).upper()
                    data['sluchawki'] = cell_data(
                        row[col_names['Słuchawki']].value)
                    data['zlacze'] = cell_data(row[col_names['Złącze']].value)
                    data['przejsciowka'] = cell_data(
                        row[col_names['Przejściówka']].value)
                    data['mysz'] = cell_data(row[col_names['Mysz']].value)
                    data['torba'] = cell_data(row[col_names['Torba']].value)
                    data['modem'] = cell_data(row[col_names['Modem']].value)
                    notatki_wypozyczenie = row[col_names['Notatki wypożyczenie']].value
                    if notatki_wypozyczenie:
                        data['notatki_wypozyczenie'] = f"""{notatki_wypozyczenie}
                      (Sprzęt udostępniony z pliku - {datetime.now().strftime('%Y-%m-%d')})
                      """
                    else:
                        data['notatki_wypozyczenie'] = f"Sprzęt udostępniony z pliku - {datetime.now().strftime('%Y-%m-%d')}"
                    data['rented_status'] = True
                    data['rent_date'] = rent_date.strftime(
                        "%Y-%m-%d") if rent_date else ""
                    data['last_updated'] = local_time
                    data['who_rented'] = who_rented if who_rented else ""
                db_items.insert_one(data)
                if data['rented_status']:
                    data_for_history = data.copy()
                    data_for_history['modyfikacja'] = 'Wypożyczony'
                    data_for_history['who_modified'] = current_user.login
                    db_history.insert_one(data_for_history)
                    update_for_cron("sm_history")
                data_table.append(data)
                update_for_cron("sm_items")
                update_for_cron("sm_history")
            else:
                if barcode != None:
                    db_barcodes.append(barcode)

    if data_table and db_barcodes:
        flash(
            f"""Pomyślnie dodano {len(data_table)} rekordów do bazy danych.\n
            Barcode\'y już istniejące w bazie danych: {db_barcodes}'""", 'success')
    elif data_table:
        flash(
            f'Pomyślnie dodano {len(data_table)} rekordów do bazy danych', 'success')
    elif db_barcodes:
        flash(
            f"""Nie dodano żadnych nowych rekordów do bazy danych.\n
            Barcode\'y już istniejące w bazie danych: {db_barcodes}""", 'warning')
    else:
        flash(
            f'Wystąpił błąd! Nie dodano żadnych nowych rekordów do bazy danych.', 'error')
    return db_barcodes


@hardware.route('/add_file', methods=['GET', 'POST'])
@login_required
def add_file():
    form = AddHardwareFromField()
    if request.method == 'POST':
        file = form.plik.data
        go_through_file(file)
        return redirect(url_for("hardware.add"))
    return render_template('add_file.html', form=form)


@hardware.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    hardware_form = AddHardware()
    try:
        if request.method == 'POST':
            barcode_number = hardware_form.barcode.data.strip()
            if check_existing_data(barcode_number, 'barcode', db_items):
                local_time = datetime.now(
                    local_tz).strftime("%Y-%m-%d %H:%M:%S")
                stanowisko_data = data_handler(
                    hardware_form.stanowisko.data,
                    hardware_form.nowy_stanowisko.data,
                    'stanowisko'
                )
                typ_data = data_handler(
                    hardware_form.typ.data,
                    hardware_form.nowy_typ.data,
                    'type'
                )
                marka_data = data_handler(
                    hardware_form.marka.data,
                    hardware_form.nowa_marka.data,
                    'marka'
                )
                model_data = data_handler(
                    hardware_form.model.data,
                    hardware_form.nowy_model.data,
                    'model'
                )
                system_data = data_handler(
                    hardware_form.system.data,
                    hardware_form.nowy_system.data,
                    'system'
                )
                mpk_data = data_handler(
                    hardware_form.mpk.data,
                    hardware_form.nowy_mpk.data,
                    'mpk'
                )
                projekt_data = data_handler(
                    hardware_form.projekt.data,
                    hardware_form.nowy_projekt.data,
                    'projekt'
                )
                lokalizacja_data = data_handler(
                    hardware_form.lokalizacja.data,
                    hardware_form.nowa_lokalizacja.data,
                    'lokalizacja'
                )
                opis_szkod = hardware_form.opis_uszkodzenia.data
                data_to_send = {
                    'barcode': barcode_number,
                    'typ': typ_data,
                    'marka': marka_data,
                    'model': model_data,
                    'mpk': mpk_data,
                    'stan': hardware_form.stan.data,
                    'rented_status': False,
                    'upload_date': local_time,
                    'last_updated': local_time,
                    'notatki': hardware_form.notatki.data if hardware_form.notatki.data else "",
                    'opis_uszkodzenia': opis_szkod if opis_szkod else "",
                    'adder': current_user.login,
                }
                if stanowisko_data:
                    data_to_send['stanowisko'] = stanowisko_data
                if system_data:
                    data_to_send['system'] = system_data
                if hardware_form.bitlocker.data:
                    data_to_send['bitlocker'] = hardware_form.bitlocker.data
                if hardware_form.serial.data:
                    data_to_send['serial'] = hardware_form.serial.data
                if hardware_form.identyfikator.data:
                    data_to_send['identyfikator'] = hardware_form.identyfikator.data
                if hardware_form.klucz_odzyskiwania.data:
                    data_to_send['klucz_odzyskiwania'] = hardware_form.klucz_odzyskiwania.data
                data_for_history = data_to_send.copy()
                data_for_history['modyfikacja'] = 'Stworzony'
                data_for_history['who_modified'] = data_to_send['adder']
                db_history.insert_one(data_for_history)
                data_for_history = None
                if hardware_form.login.data != None and hardware_form.login.data != "":
                    try:
                        data_to_send['rented_status'] = True
                        data_to_send['login'] = hardware_form.login.data
                        data_to_send['mocarz_id'] = hardware_form.mocarz_id.data
                        data_to_send['projekt'] = projekt_data
                        data_to_send['lokalizacja'] = lokalizacja_data
                        data_to_send['karta_zblizeniowa'] = hardware_form.karta_zblizeniowa.data
                        data_to_send['sluchawki'] = hardware_form.sluchawki.data
                        data_to_send['zlacze'] = hardware_form.zlacze.data
                        data_to_send['przejsciowka'] = hardware_form.przejsciowka.data
                        data_to_send['mysz'] = hardware_form.mysz.data
                        data_to_send['torba'] = hardware_form.torba.data
                        data_to_send['modem'] = hardware_form.modem.data
                        data_to_send['notatki_wypozyczenie'] = hardware_form.notatki_wypozyczenie.data
                        data_to_send['who_rented'] = current_user.login
                        data_to_send['rent_date'] = local_time
                    except Exception as e:
                        print('Rent error: ', e)
                db_items.insert_one(data_to_send)
                if data_to_send['rented_status'] == True:
                    data_for_history = data_to_send.copy()
                    data_for_history['modyfikacja'] = 'Wypożyczony'
                    data_for_history['who_modified'] = data_to_send['who_rented']
                    db_history.insert_one(data_for_history)
                if hardware_form.stanowisko.data != '':
                    db_stanowiska.update_one({'stanowisko': hardware_form.stanowisko.data}, {"$push": {
                        'assigned_barcodes': hardware_form.barcode.data,
                    },
                    }, upsert=True)
                update_for_cron('sm_items')
                update_for_cron('sm_history')
                flash('Sprzęt dodany pomyślnie', category='success')
                return (redirect(url_for('hardware.add')))
            else:
                flash('Taki barcode już istnieje', category='error')
                return (redirect(url_for('hardware.add')))
        else:
            collection = db_collection.find_one({"_id": "main"})
            hardware_form.stanowisko.choices = sort_and_assign(
                collection['stanowisko'], False) if check_if_exists('stanowisko') else []
            hardware_form.typ.choices = sort_and_assign(
                collection['type']) if check_if_exists('type') else []
            hardware_form.marka.choices = sort_and_assign(
                collection['marka']) if check_if_exists('marka') else []
            hardware_form.model.choices = sort_and_assign(
                collection['model']) if check_if_exists('model') else []
            hardware_form.system.choices = sort_and_assign(
                collection['system'], False) if check_if_exists('system') else []
            hardware_form.mpk.choices = sort_and_assign(
                collection['mpk']) if check_if_exists('mpk') else []
            hardware_form.projekt.choices = sort_and_assign(
                collection['projekt']) if check_if_exists('projekt') else []
            hardware_form.lokalizacja.choices = sort_and_assign(
                collection['lokalizacja']) if check_if_exists('lokalizacja') else []
            return render_template('add_items.html',
                                   header_text="Dodaj",
                                   form=hardware_form,
                                   edit=False,
                                   hardware_data=False,
                                   show_rent_hardware=False,
                                   return_to="/")
    except Exception as e:
        print(str(e))


@hardware.route('/edit/<route>/<id>', methods=['GET', 'POST'])
@login_required
def edit(route, id):
    form = AddHardware()
    item_data = db_items.find_one({'_id': ObjectId(id)}, {'_id': 0,
                                                          'rented_status': 0,
                                                          'upload_date': 0,
                                                          'last_updated': 0,
                                                          'rent_date': 0,
                                                          'adder': 0,
                                                          'who_rented': 0,
                                                          'kartoteka': 0,
                                                          'bitlocker1': 0,
                                                          'bitlocker2': 0,
                                                          'mpk': 0,
                                                          'notes': 0})

    barcode_exists = db_items.find_one(
        {'_id': ObjectId(id), 'barcode': {"$exists": True}})
    if barcode_exists:
        old_barcode = barcode_exists['barcode']
    else:
        old_barcode = None
    if request.method == 'POST':
        def update_db():
            kartoteka_exists = db_items.find_one(
                {'_id': ObjectId(id), 'kartoteka': {"$exists": True}})
            if kartoteka_exists:
                existing_kartoteka = db_paperwork.find_one(
                    {'kartoteka': kartoteka_exists['kartoteka']})
            else:
                existing_kartoteka = None
            update_data = {
                'barcode': form.barcode.data,
                'stanowisko': data_handler(form.stanowisko.data, form.nowy_stanowisko.data, "stanowisko"),
                'typ': data_handler(form.typ.data, form.nowy_typ.data, "type"),
                'marka': data_handler(form.marka.data, form.nowa_marka.data, "marka"),
                'model': data_handler(form.model.data, form.nowy_model.data, "model"),
                'system': data_handler(form.system.data, form.nowy_system.data, "system"),
                'mpk': data_handler(form.mpk.data, form.nowy_mpk.data, "mpk"),
                'stan': form.stan.data,
                'bitlocker': form.bitlocker.data,
                'serial': form.serial.data,
                'identyfikator': form.identyfikator.data,
                'klucz_odzyskiwania': form.klucz_odzyskiwania.data,
                'notatki': form.notatki.data,
                'opis_uszkodzenia': form.opis_uszkodzenia.data
            }
            db_items.update_one({'_id': ObjectId(id)}, {
                                '$set': update_data}, upsert=True)
            update_for_cron('sm_items')
            update_history('_id', ObjectId(id), 'Edytowany')
            if existing_kartoteka:
                for barcode in existing_kartoteka['przypisane_barcodes']:
                    if barcode == old_barcode:
                        existing_kartoteka['przypisane_barcodes'].remove(
                            barcode)
                        existing_kartoteka['przypisane_barcodes'].append(
                            form.barcode.data)
                db_paperwork.update_one({'kartoteka': existing_kartoteka['kartoteka']}, {'$set': {
                                        'przypisane_barcodes': existing_kartoteka['przypisane_barcodes']}}, upsert=True)
                update_for_cron("sm_paperwork")
            flash('Zmiany naniesione pomyślnie', category='success')
        new_barcode = form.barcode.data
        if new_barcode != old_barcode:
            existing_id = db_items.find_one({'barcode': new_barcode})
            if not existing_id:
                update_db()
                return (redirect(url_for('hardware.show_info', data='present', id=id)))
            else:
                flash('Taki barcode już istnieje', category='error')
                return (redirect(url_for('hardware.edit', route=route, id=id)))
        else:
            update_db()
            return (redirect(url_for('hardware.show_info', data='present', id=id)))
    else:
        for key, value in item_data.items():
            form[key].data = value
        collection = db_collection.find_one({"_id": "main"})
        form.stanowisko.choices = sort_and_assign(
            collection['stanowisko'], False) if check_if_exists('stanowisko') else []
        form.typ.choices = sort_and_assign(
            collection['type']) if check_if_exists('type') else []
        form.marka.choices = sort_and_assign(
            collection['marka']) if check_if_exists('marka') else []
        form.model.choices = sort_and_assign(
            collection['model']) if check_if_exists('model') else []
        form.system.choices = sort_and_assign(
            collection['system'], False) if check_if_exists('system') else []
        form.mpk.choices = sort_and_assign(
            collection['mpk']) if check_if_exists('mpk') else []
        form.projekt.choices = sort_and_assign(
            collection['projekt']) if check_if_exists('projekt') else []
        form.lokalizacja.choices = sort_and_assign(
            collection['lokalizacja']) if check_if_exists('lokalizacja') else []
        if route == 'all':
            return_route = "/hardware/all"
        else:
            return_route = f"/hardware/show_info/present/{id}"
        return render_template('add_items.html',
                               header_text="Edytuj",
                               data=item_data,
                               hardware_data=False,
                               edit=True,
                               form=form,
                               return_to=return_route)


@ hardware.route('/rent/<route>/<barcode>', methods=['GET', 'POST'])
@ login_required
def rent(route, barcode):
    local_time = datetime.now(local_tz).strftime("%Y-%m-%d %H:%M:%S")
    hardware_form = AddHardware()
    hardware_data = db_items.find_one({'barcode': barcode}, {
        '_id': 0,
        'login': 1,
        'mocarz_id': 1,
        'projekt': 1,
        'lokalizacja': 1,
        'karta_zblizeniowa': 1,
        'sluchawki': 1,
        'zlacze': 1,
        'przejsciowka': 1,
        'mysz': 1,
        'modem': 1
    })
    return_route = adjust_return_route(route, barcode)
    if request.method == 'POST':
        projekt_data = data_handler(
            hardware_form.projekt.data,
            hardware_form.nowy_projekt.data,
            'projekt'
        )
        lokalizacja_data = data_handler(
            hardware_form.lokalizacja.data,
            hardware_form.nowa_lokalizacja.data,
            'lokalizacja'
        )
        db_items.update_one({'barcode': barcode}, {"$set": {
            'login': hardware_form.login.data,
            'mocarz_id': hardware_form.mocarz_id.data,
            'projekt': projekt_data,
            'lokalizacja': lokalizacja_data,
            'karta_zblizeniowa': hardware_form.karta_zblizeniowa.data,
            'sluchawki': hardware_form.sluchawki.data,
            'zlacze': hardware_form.zlacze.data,
            'przejsciowka': hardware_form.przejsciowka.data,
            'mysz': hardware_form.mysz.data,
            'torba': hardware_form.torba.data,
            'modem': hardware_form.modem.data,
            'notatki_wypozyczenie': hardware_form.notatki_wypozyczenie.data,
            'rented_status': True,
            'rent_date': local_time,
            'last_updated': local_time,
            'who_rented': current_user.login
        }
        }
        )
        update_for_cron("sm_items")
        update_history('barcode', barcode, 'Wypożyczony')
        if route == 'all':
            return (redirect(url_for('hardware.see_all')))
        else:
            return (redirect(return_route))
    else:
        collection = db_collection.find_one({"_id": "main"})
        hardware_form.projekt.choices = sort_and_assign(
            collection['projekt']) if check_if_exists('projekt') else []
        hardware_form.lokalizacja.choices = sort_and_assign(
            collection['lokalizacja']) if check_if_exists('lokalizacja') else []
        hardware_form.sluchawki.choices = sort_and_assign(
            collection['sluchawki'], mandatory=False, sluchawki_zlacze=True) if check_if_exists('sluchawki') else ['Nie dotyczy']
        hardware_form.zlacze.choices = sort_and_assign(
            collection['zlacze'], mandatory=False, sluchawki_zlacze=True) if check_if_exists('sluchawki') else ['Nie dotyczy']
        if route == 'edit':
            header_text = "Edytuj"
            for key, value in hardware_data.items():
                hardware_form[key].data = value
        else:
            header_text = "Udostępnij"
        return render_template('add_items.html',
                               udostepnienie=True,
                               header_text=header_text,
                               form=hardware_form,
                               return_to=return_route
                               )


@ hardware.route('/see_history/<id>/<barcode>')
@ login_required
def see_history(id, barcode):
    hardware_data_history = db_history.find({'barcode': barcode})
    look_for_adder_and_upload_date = list(hardware_data_history.clone())
    creator_value = None
    date_value = None
    for each in look_for_adder_and_upload_date:
        if creator_value == None:
            creator_value = each.get('adder', None)
        else:
            break
    for each in look_for_adder_and_upload_date:
        if date_value == None:
            date_value = each.get('upload_date', None)
        else:
            break
    creation_date = date_value if date_value else "N/A"
    creator = creator_value if creator_value else "N/A"
    return render_template("hardware_history.html",
                           creator=creator,
                           barcode=barcode,
                           date=creation_date,
                           history=hardware_data_history,
                           return_to=f"/hardware/show_info/present/{id}")


@ hardware.route('/return/<route>/<barcode>', methods=['GET', 'POST'])
@ login_required
def return_hardware(route, barcode):
    form = ReturnHardware()
    hardware_data = db_items.find_one({'barcode': barcode})
    return_route = adjust_return_route(route, barcode)
    if request.method == "POST":
        stan = form.stan.data
        opis_szkod = form.opis_uszkodzenia.data
        dodatkowe_uwagi = form.dodatkowe_uwagi.data
        db_items.update_one({'barcode': barcode}, {
            "$set": {
                'rented_status': False,
                'last_updated': datetime.now(local_tz).strftime("%Y-%m-%d %H:%M:%S"),
                'notatki': dodatkowe_uwagi,
                'opis_uszkodzenia': opis_szkod,
                'stan': stan,
            },
            "$unset": {
                'mocarz_id': "",
                'projekt': "",
                'lokalizacja': "",
                'rent_date': "",
                'who_rented': "",
                'notatki_wypozyczenie': "",
                'karta_zblizeniowa': "",
                'sluchawki': "",
                'zlacze': "",
                'przejsciowka': "",
                'mysz': "",
                'torba': "",
                'modem': "",
            }
        }, upsert=True
        )
        update_for_cron("sm_items")
        update_history('barcode', barcode, "Zwrócony")
        if route == 'all':
            return (redirect(url_for('hardware.see_all')))
        else:
            return (redirect(return_route))
    else:
        form['stan'].data = hardware_data['stan']
        if hardware_data['notatki']:
            form['dodatkowe_uwagi'].data = hardware_data['notatki']
        if hardware_data['opis_uszkodzenia']:
            form['opis_uszkodzenia'].data = hardware_data['opis_uszkodzenia']
        return render_template("return_hardware.html",
                               hardware_data=hardware_data,
                               paperwork_data=None,
                               being_returned=True,
                               form=form,
                               return_to=return_route)


@ hardware.route('/show_info/<data>/<id>')
@ login_required
def show_info(data, id):
    if data == 'present':
        hardware_data = db_items.find_one({'_id': ObjectId(id)})
        hide_buttons = False
        return_route = "/hardware/all"
        check_kartoteka = db_items.find_one(
            {"$and": [{'_id': ObjectId(id)}, {'kartoteka': {"$exists": True}}]})
        if check_kartoteka == None:
            paperwork_data = None
        else:
            paperwork_data = db_paperwork.find_one(
                {'kartoteka': check_kartoteka['kartoteka']})
    else:
        hardware_data = db_history.find_one({'_id': ObjectId(id)})
        return_data = db_items.find_one({'barcode': hardware_data['barcode']})
        return_barcode = str(hardware_data['barcode'])
        return_id = str(return_data['_id'])
        hide_buttons = True
        return_route = f"/hardware/see_history/{return_id}/{return_barcode}"
        check_kartoteka = db_history.find_one(
            {"$and": [{'_id': ObjectId(id)}, {'kartoteka': {"$exists": True}}]})
        if check_kartoteka == None:
            paperwork_data = None
        else:
            paperwork_data = check_kartoteka
    return render_template("information_page.html",
                           hardware_data=hardware_data,
                           paperwork_data=paperwork_data,
                           return_to=return_route,
                           hide_buttons=hide_buttons
                           )


@ hardware.route('/details')
@ login_required
def details():
    return render_template('hardware_details.html')


@ hardware.route('/all', methods=["GET", "POST"])
@ login_required
def see_all():
    if request.method == 'POST':
        return redirect(url_for('main.index'))
    else:
        if current_user.dostep == "User":
            all_items = db_items.find({'mpk': {"$in": current_user.mpk}})
        else:
            all_items = db_items.find({})
        return render_template('all_items.html', items=all_items, data=navbar_select_data, see_hardware=True, see_all=True)


def get_hardware_data_by_status(search_val, search_bool):
    if search_val == 'barcode':
        if current_user.dostep == "User":
            free_items = db_items.find(
                {search_val: {"$exists": search_bool}, 'mpk': {"$in": current_user.mpk}})
        else:
            free_items = db_items.find({search_val: {"$exists": search_bool}})
    else:
        if current_user.dostep == "User":
            free_items = db_items.find(
                {search_val: search_bool, 'mpk': {"$in": current_user.mpk}})
        else:
            free_items = db_items.find({search_val: search_bool})
    return free_items


@ hardware.route('/get_data/<parametr>')
@ login_required
def get_data(parametr):
    if parametr == "no-barcode":
        free_items = get_hardware_data_by_status('barcode', False)
    elif parametr == "barcode":
        free_items = get_hardware_data_by_status('barcode', True)
    elif parametr == "not-rented":
        free_items = get_hardware_data_by_status('rented_status', False)
    elif parametr == "rented":
        free_items = get_hardware_data_by_status('rented_status', True)
    return render_template('all_items.html', items=free_items, data=navbar_select_data, see_hardware=True, see_all=True)


@ hardware.route('/delete/<route>/<id>', methods=["GET", "POST"])
def delete(route, id):
    if request.method == 'POST':
        kartoteka_attached = db_items.find_one(
            {'_id': ObjectId(id), 'kartoteka': {"$exists": True}})
        if kartoteka_attached:
            kartoteka = db_paperwork.find_one(
                {'kartoteka': kartoteka_attached['kartoteka']})
            for barcode in kartoteka['przypisane_barcodes']:
                if kartoteka_attached['barcode'] == barcode:
                    db_paperwork.update_one({'kartoteka': kartoteka_attached['kartoteka']}, {
                                            "$pull": {'przypisane_barcodes': barcode}})
                    update_for_cron("sm_paperwork")
        update_history('_id', ObjectId(id), "Usunięty")
        db_items.delete_one({'_id': ObjectId(id)})
        update_for_cron("sm_items")
        return (redirect(url_for('hardware.see_all')))
    else:
        if route == 'all':
            return_route = "/hardware/all"
        else:
            return_route = f"/hardware/show_info/present/{id}"
        return render_template("confirmation.html", id=id, return_to=return_route)
