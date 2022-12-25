from datetime import datetime
from flask import Blueprint, flash, render_template, redirect, url_for, request
from website.extensions import *
from website.constants import *
from ..forms import AddPaperwork, AddHardwareFromField
from flask_login import login_required, current_user
import openpyxl
from bson import ObjectId

paperwork = Blueprint('paperwork', __name__)

navbar_select_data = [('', ''), ('all', 'Wszystkie'), ('no-faktura', 'Brak faktury'), (
    'faktura', 'Z fakturą'), ('no-barcode', 'Brak barcode\'u'), ('barcode', 'Z barcodem')]


def check_existing_records(kartoteka):
    existing_record = db_paperwork.find_one({'kartoteka': kartoteka})
    if existing_record == None:
        return True
    else:
        return False


def copy_and_update_item(barcode, msg, faktury, kartoteka_typ,
                         mpk, notatki, data_faktury):
    updated_item = db_items.find_one(
        {'barcode': barcode}, {'_id': 0})
    data_for_history = updated_item.copy()
    data_for_history['przypisane_faktury'] = faktury
    data_for_history['kartoteka_typ'] = kartoteka_typ
    data_for_history['kartoteka_mpk'] = mpk
    data_for_history['kartoteka_notatki'] = notatki
    data_for_history['data_faktury'] = data_faktury
    data_for_history['modyfikacja'] = msg
    data_for_history['who_modified'] = current_user.login
    data_for_history['last_updated'] = datetime.now(
        tz=local_tz).strftime("%Y-%m-%d %H:%M:%S")
    db_history.insert_one(data_for_history)
    update_for_cron("sm_history")

def go_through_file(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file)
    ws_dokumentacja = wb['Dokumentacja']
    dokumentacja_table = []
    for row in range(2, ws_dokumentacja.max_row+1):
        if row:
            kartoteka = ws_dokumentacja.cell(row, 2).value
            if check_existing_records(kartoteka):
                barcode_row = ws_dokumentacja.cell(row, 1).value
                faktury_row = ws_dokumentacja.cell(row, 3).value
                notatki_dokumentacja = ws_dokumentacja.cell(row, 7).value
                data = {
                    'kartoteka': ws_dokumentacja.cell(row, 2).value if ws_dokumentacja.cell(row, 2).value != None else "N/A",
                    'kartoteka_typ': ws_dokumentacja.cell(row, 4).value if ws_dokumentacja.cell(row, 4).value != None else "N/A",
                    'kartoteka_mpk': ws_dokumentacja.cell(row, 5).value if ws_dokumentacja.cell(row, 5).value != None else "N/A",
                    'data_faktury': ws_dokumentacja.cell(row, 6).value if ws_dokumentacja.cell(row, 6).value != None else "",
                    'kartoteka_notatki': f"{notatki_dokumentacja if notatki_dokumentacja != None else ''}\n Kartoteka dodana z pliku - {datetime.now().strftime('%Y-%m-%d')}"
                }
                if barcode_row:
                    barcodes = []
                    barcodes_to_strip = barcode_row.split(',')
                    for each in barcodes_to_strip:
                        barcodes.append(each.strip())
                    for barcode in barcodes:
                        db_items.update_one({'barcode': barcode}, {
                                            '$set': {'kartoteka': kartoteka}})
                    update_for_cron("sm_items")
                    data['przypisane_barcodes'] = barcodes
                if faktury_row:
                    faktury = []
                    faktury_to_strip = faktury_row.split(',')
                    for each in faktury_to_strip:
                        faktury.append(each.strip())
                    data['przypisane_faktury'] = faktury

                dokumentacja_table.append(data)
            else:
                print('Boof')

    if dokumentacja_table:
        db_paperwork.insert_many(dokumentacja_table)


@paperwork.route('/add_file', methods=['GET', 'POST'])
@login_required
def add_file():
    form = AddHardwareFromField()
    if request.method == 'POST':
        file = form.plik.data
        go_through_file(file)
        return redirect(url_for("paperwork.add"))
    return render_template('add_file.html', form=form)


@paperwork.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    form = AddPaperwork()
    free_items = db_items.find(
        {'kartoteka': {"$exists": False}, 'barcode': {"$exists": True}}).sort('barcode', 1)
    if request.method == 'POST':
        try:
            if check_existing_records(form.kartoteka.data):
                mpk_data = data_handler(
                    form.mpk.data, form.nowy_mpk.data, "mpk")
                faktury = form.faktury.data.split(',')
                barcodes = request.form.getlist('barcodes')
                data = {
                    'kartoteka': form.kartoteka.data,
                    'przypisane_faktury': faktury,
                    'kartoteka_typ': form.kartoteka_typ.data,
                    'kartoteka_mpk': mpk_data,
                    'kartoteka_notatki': form.notatki.data,
                    'data_dodania': datetime.now(local_tz).strftime("%Y-%m-%d %H:%M:%S"),
                }
                if form.data_przyjecia.data:
                    data['data_faktury'] = form.data_przyjecia.data.strftime(
                        "%Y-%m-%d")
                else:
                    data['data_faktury'] = ""
                if len(barcodes) > 0:
                    data['przypisane_barcodes'] = barcodes
                    for each in data['przypisane_barcodes']:
                        db_items.update_one({'barcode': each}, {"$set": {
                            'kartoteka': form.kartoteka.data,
                        }})
                        copy_and_update_item(
                            each,
                            'Dodano kartotekę',
                            faktury,
                            form.kartoteka_typ.data,
                            mpk_data,
                            form.notatki.data,
                            data['data_faktury'])
                    update_for_cron("sm_items")
                db_paperwork.insert_one(data)
                update_for_cron("sm_paperwork")
                flash('Dodano kartotekę', 'success')
                return redirect(url_for('paperwork.add'))
            else:
                flash('Kartoteka istnieje', 'error')
                return redirect(url_for('paperwork.add'))
        except Exception as e:
            print(str(e))
    else:
        collection = db_collection.find_one({})
        form.mpk.choices = sort_and_assign(
            collection['mpk'], False) if check_if_exists('mpk') else []
        return render_template('add_paperwork.html',
                               header_text="Dodaj",
                               edit=False, form=form,
                               available_barcodes=free_items,
                               return_to="/")


@paperwork.route('/edit/<id>', methods=['GET', 'POST'])
def edit(id):
    try:
        find_barcodes = db_paperwork.find_one({'_id': ObjectId(id)}, {
            'przypisane_barcodes': 1, '_id': 0})
        if find_barcodes:
            barcodes_from_db = find_barcodes['przypisane_barcodes']
        else:
            barcodes_from_db = []
        barcodes_to_delete = []
        all_items = []
        form = AddPaperwork()
        free_items = db_items.find(
            {'kartoteka': {"$exists": False}, 'barcode': {"$exists": True}})
        for each in free_items:
            all_items.append(each)
        for each in barcodes_from_db:
            item = db_items.find_one({'barcode': each}, {'_id': 0})
            all_items.append(item)
        sorted_items = sorted(all_items, key=lambda k: k['barcode'])
        if request.method == 'POST':
            mpk_data = data_handler(form.mpk.data, form.nowy_mpk.data, "mpk")
            faktury = form.faktury.data.split(',')
            barcodes_from_select = sorted(request.form.getlist('barcodes'))
            if form.data_przyjecia.data:
                data_faktury = form.data_przyjecia.data.strftime("%Y-%m-%d")
            else:
                data_faktury = ""
            db_paperwork.update_one({'_id': ObjectId(id)}, {'$set': {
                'kartoteka': form.kartoteka.data,
                'kartoteka_typ': form.kartoteka_typ.data,
                'kartoteka_mpk': mpk_data,
                'data_faktury': data_faktury,
                'przypisane_barcodes': barcodes_from_select,
                'przypisane_faktury': faktury,
                'kartoteka_notatki': form.notatki.data,
                'update_date': datetime.now(
                    local_tz).strftime("%Y-%m-%d %H:%M:%S")
            }}, upsert=True)
            update_for_cron("sm_paperwork")
            for each in barcodes_from_select:
                if not db_items.find_one({'barcode': each, 'kartoteka': form.kartoteka.data}):
                    db_items.update_one({'barcode': each}, {"$set": {
                        'kartoteka': form.kartoteka.data
                    }})
                    update_for_cron("sm_items")
                    copy_and_update_item(
                        each,
                        'Dodano kartotekę',
                        faktury,
                        form.kartoteka_typ.data,
                        mpk_data,
                        form.notatki.data,
                        data_faktury)

            for barcode in barcodes_from_db:
                if barcode not in barcodes_from_select:
                    barcodes_to_delete.append(barcode)
            if len(barcodes_to_delete) > 0:
                for barcode in barcodes_to_delete:
                    db_items.update_one({'barcode': barcode}, {"$unset": {
                        'kartoteka': ""
                    }})
                    updated_item = db_items.find_one(
                        {'barcode': barcode}, {'_id': 0})
                    data_for_history = updated_item.copy()
                    data_for_history['modyfikacja'] = 'Usunięto kartotekę'
                    data_for_history['who_modified'] = current_user.login
                    data_for_history['last_updated'] = datetime.now(
                        tz=local_tz).strftime("%Y-%m-%d %H:%M:%S")
                    db_history.insert_one(data_for_history)
                update_for_cron("sm_items")
                update_for_cron("sm_history")

            flash('Zaktualizowano kartotekę', 'success')
            return (redirect(url_for('paperwork.see_all')))
        else:
            rest = {}
            collection = db_collection.find_one({})
            form.mpk.choices = sort_and_assign(
                collection['mpk'], False) if check_if_exists('mpk') else []
            data = db_paperwork.find_one({'_id': ObjectId(id)}, {
                '_id': 0, 'update_date': 0})
            for key, value in data.items():
                if key == 'przypisane_barcodes':
                    barcodes_from_db = value
                    form['barcodes_form'].data = list_to_str(value)
                elif key == 'przypisane_faktury':
                    form['faktury'].data = list_to_str(value)
                elif key == 'data_faktury':
                    if value == '' or value == None:
                        form['data_przyjecia'].data = ''
                    else:
                        form['data_przyjecia'].data = datetime.strptime(
                            value, '%Y-%m-%d')
                elif key == 'kartoteka_mpk':
                    form['mpk'].data = value
                elif key == 'kartoteka_notatki':
                    form['notatki'].data = value
                else:
                    rest[key] = value

            for key, value in rest.items():
                form[key].data = value
    except Exception as e:
        print(e)
        flash('Błąd: ' + str(e), 'error')
    return render_template('add_paperwork.html', header_text="Edytuj", form=form,
                           available_barcodes=free_items,
                           all_barcodes=sorted_items,
                           edit=True,
                           return_to="/paperwork/all")


@ paperwork.route('/all')
def see_all():
    all_items = db_paperwork.find({})
    return render_template('all_papers.html', items=all_items, data=navbar_select_data, see_all=True)


@ paperwork.route('/delete/<id>', methods=["GET", "POST"])
def delete(id):
    if request.method == "POST":
        all_items = db_paperwork.find({})
        db_has_barcodes = db_paperwork.find_one(
            {'_id': ObjectId(id), 'przypisane_barcodes': {"$exists": True}})
        if db_has_barcodes:
            barcodes = db_has_barcodes['przypisane_barcodes']
            for barcode in barcodes:
                db_items.update_one({'barcode': barcode}, {"$unset": {
                    'kartoteka': ""
                }})
            update_for_cron("sm_items")
        db_paperwork.delete_one({'_id': ObjectId(id)})
        update_for_cron("sm_paperwork")
        return redirect(url_for('paperwork.see_all'))
    else:
        return render_template("confirmation.html", id=id, return_to="/paperwork/all")


@ paperwork.route('/get_data/<parametr>')
@ login_required
def get_data(parametr):
    all_items = db_paperwork.find({})
    if parametr == "no-faktura":
        all_items = db_paperwork.find(
            {'przypisane_faktury': {"$exists": False}})
    elif parametr == "faktura":
        all_items = db_paperwork.find(
            {'przypisane_faktury': {"$exists": True}})
    elif parametr == "no-barcode":
        all_items = db_paperwork.find(
            {'przypisane_barcodes':  {"$exists": False}})
    elif parametr == "barcode":
        all_items = db_paperwork.find(
            {'przypisane_barcodes':  {"$exists": False}})
    return render_template('all_papers.html', items=all_items, data=navbar_select_data, see_all=True)
